from openpyxl import load_workbook
from openpyxl import Workbook 
from tkinter import *  
from tkinter import ttk
import tkinter as tk
from tkinter.font import Font
from tkinter import messagebox
from PIL import ImageTk, Image
import numpy as np
import xlrd
import os
import tksheet
from tkinter import messagebox
import xlwt
import datetime
from invoice import *
import mysql.connector
''' 
#CREATION OF NEW WORKBOOK
wb = Workbook() 
wb.save("demo.xlsx")
'''

productsList,productsListIds=[],[]
idList={}
label=["Product Id", "Name", "Quantity","Cost per Unit"]
ip=''

def printDatabase(r,c):
    row_count = r
    column_count = c
    for i in range(1, row_count + 1):
        for j in range(1, column_count + 1):
            data = sheet.cell(row=i, column=j).value
            print(data, end='   ')
        print('\n')

def updateDatabaseQ(r):
    row_count = r
    for i in range(2, row_count + 1):
        #for j in range(1, column_count + 1):
        sheet.cell(row=i, column=3).value += 10


class Product:
    def __init__(self,id,name,quantity,pricePerUnit,curr_date_time):
        self.productId=id
        self.productName=name
        self.productQuantity=quantity
        self.productPricePerUnit=pricePerUnit
        self.curr_date_time=curr_date_time

class Store:
    def __init__(self,productList):
        self.productList=productList
    
    def printList(self):
        for p in self.productList:
            print(p.productId,p.productName,p.productQuantity,p.productPricePerUnit,p.curr_date_time)
    def invoi(self):
        self.invoid=1234
        invo(1234,el1,el2,el3)
        prod(productList)

    # def save_info(self):
    #     infod=Tk()
    #     infod.title("Details")
    #     
        # file = open("NIK.txt","w")
        # for p in self.productList:
        #     file.write("{: >5} {: >2} {: >5} {: >5} {: >5}".format(p.curr_date_time,p.productName,p.productQuantity,p.productPricePerUnit,(p.productQuantity*p.productPricePerUnit)))
        #     file.write("\n")
        # file.close()
        # wks.save(filename="Book1.xlsx")
        # wks.close
   
    def checkout(self):
        global e1
        total=0
        for p in productsList:
            if(p.productQuantity>=1):
                total += p.productQuantity * p.productPricePerUnit

        print("Total = ",total)

        # e1 = Entry(fram3, width=30, fg='black',font=('Arial',16,'bold'))
        # e1.grid(row=0 ,column=0,columnspan=3,sticky=N+E+S )
        # e1.insert(END,'Total Amount:')
        
 
        # e1 = Entry(fram3, width=30, fg='black',font=('Arial',16,'bold'))
        # e1.grid(row=0 ,column=3)
        # e1.insert(END,total)
        print()
        print("#####################")
        #printDatabase(row_count,column_count)
        store.printList()
        ne=Tk()
        ne.title("Bill generation")
        ne.geometry("800x700")
        
        q1=tk.StringVar()
        q2=tk.StringVar()
        q3=tk.StringVar()
        r=tk.IntVar()
        r.set("0")

        wrap1=LabelFrame(ne,text="Customer Details",height=50,font=('Arial',8))
        wrap2=LabelFrame(ne,text="Bill",height=390,font=('Arial',8))
        wrap3=LabelFrame(ne,text="Payment",height=210,font=('Arial',8))

        wrap1.pack(fill="both",expand="yes",padx=20)
        wrap2.pack(fill="both",expand="yes",padx=20,pady=5)
        wrap3.pack(fill="both",expand="yes",padx=20,pady=5)

        wrap1.grid_propagate(False)
        wrap2.grid_propagate(False)
        wrap3.grid_propagate(False)

        wrap4=LabelFrame(wrap2,height=300)
        wrap5=LabelFrame(wrap2,height=20)

        mycanva=Canvas(wrap4)
        mycanva.pack(side=LEFT)
        yscrollbar=ttk.Scrollbar(wrap4,orient="vertical",command=mycanva.yview)
        yscrollbar.pack(side=RIGHT,fill="y")
        myframe=Frame(mycanva)

        wrap4.pack(fill="both",expand="yes",padx=20)
        wrap5.pack(fill="both",expand="yes",padx=20)
        wrap4.grid_propagate(False)
        wrap5.grid_propagate(False)
        
        l1=Label(wrap1,text="Name:")
        l1.grid(row=4,column=0,padx=5,pady=8)
        l2=Label(wrap1,text="Mail Id:")
        l2.grid(row=10,column=0,padx=5,pady=8)
        l3=Label(wrap1,text="Mobile No:")
        l3.grid(row=4,column=2,padx=5,pady=8)
        el1=Entry(wrap1,width=30,textvariable=q1)
        el1.grid(row=4,column=1,padx=5,pady=8,sticky=W)
        el2=Entry(wrap1,width=60,textvariable=q2)
        el2.grid(row=10,column=1,padx=5,pady=8)
        el3=Entry(wrap1,width=40,textvariable=q3)
        el3.grid(row=4,column=3,padx=5,pady=8)
        userName=el1.get()
        userMailid=el2.get()
        userPhoneNo=el3.get()
        def disp():
            global userName,userMailid,userPhoneNo
            print(userName,userMailid,userPhoneNo)
        submit=Button(wrap1,text="SUBMIT",command=disp)
        submit.grid(row=10,column=3,padx=5,pady=8)
        
        head1=Label(wrap4,text="Product ID",width=10,font=('Arial',12,'bold'))
        head1.grid(row=0,column=0)
        head2=Label(wrap4,text="Product Name",width=10,font=('Arial',12,'bold'))
        head2.grid(row=0,column=10,padx=20)
        head3=Label(wrap4,text="Price Per Unit",width=10,font=('Arial',12,'bold'))
        head3.grid(row=0,column=20,padx=20)
        head4=Label(wrap4,text="Quantity",width=10,font=('Arial',12,'bold'))
        head4.grid(row=0,column=30,padx=20)
        head5=Label(wrap4,text="Total",width=10,font=('Arial',12,'bold'))
        head5.grid(row=0,column=40)
        co=1
        for y in self.productList:
            label=Label(wrap4,text=y.productId,width=10,font=('Arial',10))
            label.grid(row=co,column=0)
            label1=Label(wrap4,text=y.productName,width=10,font=('Arial',10))
            label1.grid(row=co,column=10,padx=20)
            
            label2=Label(wrap4,text=y.productPricePerUnit,width=10,font=('Arial',10))
            label2.grid(row=co,column=20,padx=20)
            
            label3=Label(wrap4,text=y.productQuantity,width=10,font=('Arial',10))
            label3.grid(row=co,column=30,padx=20)
            
            label4=Label(wrap4,text=(y.productPricePerUnit*y.productQuantity),width=10,font=('Arial',10))
            label4.grid(row=co,column=40)
    
            co+=1
        label5=Label(wrap5,text="Final Amount:",width=20, fg='black',font=('Arial',16,'bold'))
        label5.grid(row=0,column=30,padx=10,sticky=W)
        label6=Label(wrap5,text=total,width=30, fg='black',font=('Arial',16,'bold'))
        label6.grid(row=0,column=60,padx=10,sticky=W)

        def sel():
            selection = "You selected the option "
            la.config(text = selection)

            qr=Tk()
            qr.title("QR Code 4 Payment")
            qr.geometry("300x500")
            Qcanvas = Canvas(qr, width=200, height=400)
            Qcanvas.pack()
            basewidth = 150
            Qimg = Image.open("images/image2.jpeg")
            '''
            wpercent = (basewidth / float(Qimg.size[0]))
            hsize = int((float(Qimg.size[1]) * float(wpercent)))
            Qimg = Qimg.resize((basewidth, hsize), Image.ANTIALIAS)
            '''
            Qimg = ImageTk.PhotoImage(Qimg)
            Qcanvas.create_image(10,50, anchor=NW, image=Qimg)
            qr.mainloop()

        
        
        #for (text,valu) in vals.items():
        #R1 = Radiobutton(wrap3, text="Debit", variable=r, value=1,indicator=0,background="light blue",command=lambda: sel(1))
        #R1.pack( fill=X,anchor=W)
        #R2 = Radiobutton(wrap3, text="Net banking", variable=r, value=2,indicator=0,background="light blue",command=lambda: sel(2))
        #R2.pack( fill=X,anchor=W)
        R3 = Radiobutton(wrap3, text="Proceed with Payment (UPI)", variable=r, value=3,indicator=0,background="light blue",command=sel)
        R3.pack( fill=X,anchor=W)

        done=Button(wrap3,text="Done Payment")
        done.grid(row=4,column=3,padx=5,pady=8)
        done.pack()
        back=Button(wrap3,text="Go Back")
        back.grid(row=10,column=1,padx=5,pady=8)
        back.pack()

        la=Label(wrap3)
        la.pack()
        
        ne.mainloop()
        




def createProductList(pid,status=False):
    global curr_lst
    curr_lst=[]
    if pid:
        rownum=None
        for i in range(2, row_count + 1):
            data = str(sheet.cell(row=i, column=1).value)
            #print(i,pid,data)
            if data==pid:
                rownum=i
                pid,pname,pQ,priceU=readData(rownum)
                #print('rownum',rownum)
                if pid in productsListIds:
                    for p in productsList:
                        if pid==p.productId:
                            pQ+=1

                            
                            if (sheet.cell(row=rownum, column=3).value>0):
                                sheet.cell(row=rownum, column=3).value -= 1
                                if (status):
                                    p.productQuantity+=1
                                    #p.productQuantity-=1
                                elif (not status) and p.productQuantity>0:
                                    p.productQuantity-=1
                                curr_lst=[pid,pname,p.productQuantity,priceU]
                                print('CCC',curr_lst)
                            else:
                                print(sheet.cell(row=rownum, column=2).value,"- Not Available any more")
                                curr_lst=[pid,pname,pQ,priceU]
                                print('CCC---',curr_lst)

                            
                            
                else:
                    #pid,pname,pQ,priceU=readData(rownum)
                    current_date_and_time = datetime.datetime.now()
                    current_date_and_time_string = str(current_date_and_time)
                    curr_lst=[pid,pname,pQ,priceU]
                    productsList.append(Product(pid,pname,pQ,priceU,current_date_and_time_string))
                    productsListIds.append(pid)
                    break
       
    else:
        print("enter valid pid no.")

    return curr_lst
        
        
    


def readData(rownum):
    #workbook = file
    sheet =  wb["Sheet"]
    sheet=wb.active
    pid=str(sheet.cell(row=rownum, column=1).value)
    pname=str(sheet.cell(row=rownum, column=2).value)
    if sheet.cell(row=rownum, column=3).value>0:
        pQ=1
        sheet.cell(row=rownum, column=3).value= int(sheet.cell(row=rownum, column=3).value)-1
    else:
        print(sheet.cell(row=rownum, column=2).value,"- Not Available any more")
        pQ=0
    priceU=float(sheet.cell(row=rownum, column=4).value)
    wb.save("demo.xlsx")

    return pid,pname,pQ,priceU

def add():

    messagebox.showinfo("Quantity", "Item is added")
    s.destroy()
    lst=createProductList(ip,status=True)
    print('add',ip,lst)
    v=lst[2]
    r=idList[lst[0]]-1
    createTable(row,lst)
    print(r,v)
    tableSheet.set_cell_data(r,2,value=v)
    #pass
def nothing():
    #tableSheet.set_cell_data(r,2,value=v)
    messagebox.showinfo("Quantity", "Ok, No item will be added")
    
    s.destroy()

    
def rem():
    messagebox.showinfo("Quantity", "Ok, item will be removed")
    print(ip)
    s.destroy()
    lst=createProductList(ip,status=False)
    print(lst)
    createTable(row,lst)
    print(lst)

    #v=lst[2]
    #r=idList[lst[0]]-1
    
    if v>=1:
        tableSheet.set_cell_data(r,2,value=v)
    else:
        print(idList,productsListIds)
        tableSheet.delete_row(idx = r, deselect_all = False)
        #messagebox.showinfo("Quantity", "Ok, item will be removed")
        messagebox.showwarning(title='Alert', message='All items of this category are removed')
        for ix in range(len(productsList)) :
            p=productsList[ix]
            if p.productId==ip:
                productsList.remove(productsList[ix])

        del idList[ip]
        productsListIds.remove(ip)
        print(idList,productsListIds)
        flag=False
        #print(v)
def createTable(num,listt,flag=True):
    global s,r,v
    
    #print('hello',tableSheet)
    
    if listt[0] in idList.keys():
        v=listt[2]
        r=idList[listt[0]]-1
        print(idList[listt[0]],v)

    elif(flag==False):
        
        v=listt[2]-1
        quit
    else:
        tableSheet.insert_row(listt)
        idList[listt[0]]=num
 
if __name__=="__main__":
    wb = load_workbook("demo.xlsx") 
    # Sheet is the SheetName where the data has to be entered
    sheet = wb["Sheet"]
    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column
    updateDatabaseQ(row_count)
    printDatabase(row_count,column_count)
    #GUI--------------------------------------
    #global row
    row=1
    root=Tk()  
    #App Title  
    root.title("Smart Kart Application")  
    #ttk.Label(root, text="Products List").pack()  
    #Create Panedwindow  
    panedwindow=ttk.Panedwindow(root, orient=HORIZONTAL)  
    panedwindow.pack(fill=BOTH, expand=True)  

    panedwindow1=ttk.Panedwindow(root, orient=VERTICAL)  
    panedwindow1.pack(fill=BOTH, expand=True)
    #Create Frams  
    fram1=ttk.Frame(panedwindow,width=300,height=500, relief=SUNKEN)  
    fram2=ttk.Frame(panedwindow,width=400,height=500, relief=SUNKEN)  
    fram3=ttk.Frame(panedwindow1,width=300,height=70, relief=SUNKEN)
    
    panedwindow.add(fram1, weight=1)  
    panedwindow.add(fram2, weight=4)
    panedwindow1.add(fram3, weight=4)

    canvas = Canvas(fram1, width=200, height=400)
    canvas.pack()
    e1 = Entry(fram3, width=50, fg='black',font=('Arial',16,'bold'))

    #Initial Product Details display

    table = tk.Label(fram1, text="Product Details")
    idno = Label(fram1,text= label[0]+' : ------'.upper() )
    idno.pack()
    name =  Label(fram1,text= label[1]+' : ------'.upper() )
    name.pack()
    quantity =  Label(fram1,text= label[2]+' : ------'.upper() )
    quantity.pack()
    price =  Label(fram1,text= label[3]+' : ------'.upper() )
    price.pack()

    #Initial Image display
    image_data = "images/title.png"
    img = Image.open(image_data)

    basewidth = 150

    wpercent = (basewidth / float(img.size[0]))
    hsize = int((float(img.size[1]) * float(wpercent)))
    img = img.resize((basewidth, hsize), Image.ANTIALIAS)

    img = ImageTk.PhotoImage(img)
    canvas.create_image(40, 150, anchor=NW, image=img)
    
        #pass
    def display():
        global img, image_data,table,idno,name,price,quantity,row
        global lsts,s,ip
        table.destroy()
        idno.destroy()
        name.destroy()
        quantity.destroy()
        price.destroy()
        canvas.delete('all')
        ip=entry_1.get()
        print(ip)
        image_data = "images/"+str(ip)+".jpeg"
        basewidth = 150
        img = Image.open(image_data)
        wpercent = (basewidth / float(img.size[0]))
        hsize = int((float(img.size[1]) * float(wpercent)))
        img = img.resize((basewidth, hsize), Image.ANTIALIAS)
        img = ImageTk.PhotoImage(img)
        canvas.create_image(40, 150, anchor=NW, image=img)
        print(ip)
        
        if ip in productsListIds:
            print('same prod id')
            s=Tk()
            s.title("Do you want to add again ?")
            s.geometry('400x50')
            print(ip)
            #inp=ip
            
            yes=Button(s,text='Yes',command= add)
        #yes.place(x=0,y=10)
            yes.pack(side = LEFT, padx = 30)
            no=Button(s,text='No',command = nothing)
        #no.place(x=100,y=10)
            no.pack(side = RIGHT, padx = 30)
            remove=Button(s,text='Remove',command= rem,width=100)
            remove.pack(side = RIGHT, padx = 70)
        else:
            lst=createProductList(ip)
            print(lst)
            pid,pname,pQ,priceU=str(lst[0]),str(lst[1]),1,float(lst[3])
        
            table = tk.Label(fram1, text="Product Details")
            table.pack()
        
            idno =  Label(fram1,text= label[0]+' : '+str(pid).upper() )
            idno.pack()
            name =  Label(fram1,text= label[1]+' : '+str(pname).upper() )
            name.pack()
            quantity =  Label(fram1,text= label[2]+' : '+str(pQ).upper() )
            quantity.pack()
            price =  Label(fram1,text= label[3]+' : '+str(priceU).upper() )
            price.pack()
            
            
            createTable(row,lst)
            row+=1
       

    entry_1 = tk.StringVar() 
    entry_widget_1 = tk.Entry(fram1, textvariable=entry_1,bd =2)
    entry_widget_1.place(x=50, y=50)
    display = tk.Button(fram1, text="Enter Product Id", command=display)
    display.pack()
    display.place(x=50, y=75)
    
    tableSheet = tksheet.Sheet(fram2,show_header = True,
    show_x_scrollbar = False,
    show_y_scrollbar = True,
    width = 500,
    height = 500,
    headers = label, theme = "dark blue",popup_menu_bg= "#00bfff")

    tableSheet.grid()

    
    '''
    while 1:
        inp=input('Enter=')
        if inp:
            print(inp)
            createProductList(inp)
    
        else:
            break
    '''
    print("---------------------------------------")
    store=Store(productsList)
    #store.checkout()
    print("---------------------------------------")

    check_out = tk.Button(fram3, text="check out", command=store.checkout)
    #check_out.pack()
    check_out.place(x=600, y=50)

    
    root.mainloop()
    
